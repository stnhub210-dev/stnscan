"""
스캔 자동화 통합 관리 시스템 v1.0
스캔 → AI분석 → 파일명변경 → 소송/업무분류 → HTML갱신 → 구글드라이브연동
"""
import os,re,json,sys,time,base64,io,urllib.request,urllib.parse,logging
from datetime import datetime,timedelta
from pathlib import Path

SCAN_FOLDER = os.path.join(os.path.expanduser("~"),"Desktop","스캔")
ANTHROPIC_KEY=""
GDRIVE_KEY=""
FOLDER_ID="1hT9xdNEawnkrZ78p26Kcjlm-1wdR5x-X"
KEY_FILE=os.path.join(SCAN_FOLDER,"api_key.txt")
LAWSUIT_EXCEL=os.path.join(SCAN_FOLDER,"소송목록.xlsx")
OUTPUT_HTML=os.path.join(SCAN_FOLDER,"스캔관리대장.html")
ID_CACHE=os.path.join(SCAN_FOLDER,"file_ids.json")
DONE_DB=os.path.join(SCAN_FOLDER,"processed.json")
LOG_FILE=os.path.join(SCAN_FOLDER,"auto_manager.log")
FOLDER_URL=f"https://drive.google.com/drive/folders/{FOLDER_ID}"

logging.basicConfig(level=logging.INFO,format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.FileHandler(LOG_FILE,encoding="utf-8"),logging.StreamHandler(sys.stdout)])
log=logging.getLogger(__name__)

def load_keys():
    global ANTHROPIC_KEY,GDRIVE_KEY
    if os.path.exists(KEY_FILE):
        with open(KEY_FILE,"r",encoding="utf-8") as f:
            for line in f:
                line=line.strip()
                if line.startswith("ANTHROPIC="): ANTHROPIC_KEY=line.split("=",1)[1].strip()
                elif line.startswith("GDRIVE="): GDRIVE_KEY=line.split("=",1)[1].strip()
    return bool(ANTHROPIC_KEY)

def load_done():
    try:
        if os.path.exists(DONE_DB):
            with open(DONE_DB,"r",encoding="utf-8") as f: return json.load(f)
    except: pass
    return {}

def save_done(db):
    with open(DONE_DB,"w",encoding="utf-8") as f: json.dump(db,f,ensure_ascii=False,indent=2)

def extract_text(path):
    try:
        import pdfplumber
        with pdfplumber.open(path) as pdf:
            txt="\n".join(p.extract_text() or "" for p in pdf.pages[:3])
            if txt.strip(): return txt[:3000]
    except: pass
    return ""

def pdf_to_b64(path):
    try:
        from pdf2image import convert_from_path
        imgs=convert_from_path(path,first_page=1,last_page=1,dpi=150)
        if not imgs: return None
        buf=io.BytesIO(); imgs[0].save(buf,format="JPEG",quality=80)
        return base64.b64encode(buf.getvalue()).decode()
    except: return None

def analyze_pdf(fname,text="",img_b64=None):
    content=[]
    if img_b64: content.append({"type":"image","source":{"type":"base64","media_type":"image/jpeg","data":img_b64}})
    prompt=f"""PDF 분석 후 JSON만 응답하세요.
파일명:{fname}  오늘:{datetime.now().strftime("%Y-%m-%d")}
{"내용:\n"+text[:2000] if text else ""}

{{"type":"소송 또는 업무","new_filename":"날짜_회사명_문서종류(특수문자제외,50자이내)","company":"회사/기관명","case_name":"사건명/문서명","case_number":"사건번호(없으면빈문자열)","doc_type":"문서종류","deadline":"마감일YYYY-MM-DD(없으면빈문자열)","urgency":"긴급/주의/일반","summary":"핵심내용1~2줄"}}

소송:법원서류,답변서,결정문,판결문,소장,내용증명,회생,채권,최고서,가압류
업무:계약서,공문,급여,세금,보고서,등기부,영수증"""
    content.append({"type":"text","text":prompt})
    body=json.dumps({"model":"claude-sonnet-4-5","max_tokens":600,"messages":[{"role":"user","content":content}]}).encode()
    req=urllib.request.Request("https://api.anthropic.com/v1/messages",data=body,
        headers={"Content-Type":"application/json","x-api-key":ANTHROPIC_KEY,"anthropic-version":"2023-06-01"})
    try:
        with urllib.request.urlopen(req,timeout=30) as r: data=json.loads(r.read())
        raw=re.sub(r"```json|```","",data["content"][0]["text"].strip()).strip()
        return json.loads(raw)
    except Exception as e:
        log.error(f"AI분석오류:{e}")
        return {"type":"업무","new_filename":Path(fname).stem,"company":"","case_name":fname,"case_number":"","doc_type":"기타","deadline":"","urgency":"일반","summary":"분석실패"}

def safe_rename(old_path,new_name):
    folder=os.path.dirname(old_path)
    safe=re.sub(r'[\\/:*?"<>|]',"_",new_name).strip(". ")[:60]
    new_path=os.path.join(folder,safe+".pdf")
    c=1
    while os.path.exists(new_path) and new_path!=old_path:
        new_path=os.path.join(folder,f"{safe}_{c}.pdf"); c+=1
    if new_path!=old_path: os.rename(old_path,new_path); log.info(f"파일명변경:{Path(old_path).name}→{Path(new_path).name}")
    return new_path

def update_lawsuit_excel(meta,fname):
    try:
        import openpyxl
        from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
        if not os.path.exists(LAWSUIT_EXCEL): return
        wb=openpyxl.load_workbook(LAWSUIT_EXCEL)
        ws=wb["소송목록"]
        for row in ws.iter_rows(min_row=3,values_only=True):
            if row[7] and str(row[7]).strip()==fname: return
        nr=ws.max_row+1
        for r in range(3,ws.max_row+1):
            if not ws.cell(r,8).value: nr=r; break
        s=Side(style="thin"); border=Border(left=s,right=s,top=s,bottom=s)
        def c(r,col,val,bold=False,bg=None,align="left"):
            cell=ws.cell(r,col,val)
            cell.font=Font(bold=bold,size=10,name="맑은 고딕")
            cell.alignment=Alignment(horizontal=align,vertical="center",wrap_text=True)
            cell.border=border
            if bg: cell.fill=PatternFill(fill_type="solid",fgColor=bg)
            return cell
        c(nr,1,nr-2,"","","center"); c(nr,2,meta.get("company",""))
        c(nr,3,meta.get("case_name","")); c(nr,4,meta.get("case_number",""),"","","center")
        c(nr,5,meta.get("doc_type",""),"","","center"); c(nr,6,meta.get("deadline",""),"","","center")
        ug={"긴급":"FF0000","주의":"FF9900","일반":"6c757d"}
        urgency=meta.get("urgency","일반")
        uc=c(nr,7,urgency,True,ug.get(urgency),"center"); uc.font=Font(bold=True,color="FFFFFF",size=10,name="맑은 고딕")
        c(nr,8,fname); c(nr,9,meta.get("summary",""))
        ws.row_dimensions[nr].height=28
        wb.save(LAWSUIT_EXCEL)
        log.info(f"소송목록추가:{fname}[{urgency}]")
    except Exception as e: log.error(f"소송목록업데이트오류:{e}")

def fetch_drive_ids():
    if not GDRIVE_KEY: return {}
    ids={}; pt=None
    while True:
        params={"q":f"'{FOLDER_ID}' in parents and trashed=false","fields":"nextPageToken,files(id,name)","pageSize":"1000","key":GDRIVE_KEY}
        if pt: params["pageToken"]=pt
        url="https://www.googleapis.com/drive/v3/files?"+urllib.parse.urlencode(params)
        try:
            req=urllib.request.Request(url,headers={"User-Agent":"Mozilla/5.0"})
            with urllib.request.urlopen(req,timeout=15) as r: data=json.loads(r.read())
            for f in data.get("files",[]):
                if f["name"].lower().endswith(".pdf"): ids[f["name"]]=f["id"]
            pt=data.get("nextPageToken")
            if not pt: break
        except Exception as e: log.error(f"Drive오류:{e}"); break
    if ids:
        with open(ID_CACHE,"w",encoding="utf-8") as f: json.dump(ids,f,ensure_ascii=False)
        log.info(f"DriveID {len(ids)}개")
    return ids

def load_drive_ids(force=False):
    if not force and os.path.exists(ID_CACHE):
        if time.time()-os.path.getmtime(ID_CACHE)<3600:
            with open(ID_CACHE,"r",encoding="utf-8") as f: return json.load(f)
    return fetch_drive_ids()

def load_lawsuit_dict():
    result={}
    if not os.path.exists(LAWSUIT_EXCEL): return result
    try:
        import openpyxl
        wb=openpyxl.load_workbook(LAWSUIT_EXCEL,data_only=True)
        ws=wb["소송목록"]
        for row in ws.iter_rows(min_row=3,values_only=True):
            if not row[7]: continue
            fname=str(row[7]).strip()
            result[fname]=(str(row[1] or "").strip(),str(row[2] or "").strip(),str(row[3] or "").strip(),str(row[4] or "").strip(),str(row[6] or "일반").strip())
    except Exception as e: log.error(f"소송목록읽기:{e}")
    return result

COMPANIES=[
    ("에스티엔미디어",["에스티엔미디어","에스티엔방송","STN","stn"]),
    ("에스티엔뉴스",["에스티엔뉴스","stnnews"]),
    ("에스티엔",["에스티엔등기","에스티엔 3월","에스티엔 매출","260309_에스티엔","260320_에스티엔","260401","260403_에스티엔","260406_에스티엔","260413_에스티엔","260422_에스티엔","에스티엔"]),
    ("이강영",["이강영"]),
    ("플래닛네이처미디어",["플래닛네이처미디어","플레닛네이처미디어","플래닛네이처","플레닛","플래닛부동산"]),
]

def get_date(f):
    m=re.search(r'SCAN_(\d{4})(\d{2})(\d{2})_',f)
    if m: return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    m=re.search(r'\((\d{2})\.(\d{2})\.(\d{2})\)',f)
    if m: return f"20{m.group(1)}-{m.group(2)}-{m.group(3)}"
    m=re.search(r'^(\d{2})(\d{2})(\d{2})_',f)
    if m: return f"20{m.group(1)}-{m.group(2)}-{m.group(3)}"
    m=re.search(r'(202\d)[.\-](\d{1,2})[.\-](\d{1,2})',f)
    if m: return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    return datetime.now().strftime("%Y-%m-%d")

def get_co(f):
    fn=f.replace(".pdf","")
    for n,ks in COMPANIES:
        for k in ks:
            if k in fn: return n
    for p in ["이미정","김경아","김준영","차주화","정기학","심창호","경민창","김윤석"]:
        if p in fn: return p
    return "기타"

def get_tags(f):
    fn=f.replace(".pdf",""); tags=[]
    for n,ks in COMPANIES:
        for k in ks:
            if k in fn: tags.append(n); break
    return tags

def get_dt(f):
    fn=f.replace(".pdf","")
    if "내용증명" in fn: return "✉️ 내용증명"
    if any(k in fn for k in ["계약서","협약서","양도","양수"]): return "📝 계약서"
    if any(k in fn for k in ["급여","연봉","용역비","품의","채용","이력서"]): return "👤 인사/급여"
    if any(k in fn for k in ["국세","지방세","세무소","과태료","건강보험"]): return "🧾 세금/공과"
    if any(k in fn for k in ["등기부","등기권리","사업자등록","법인인감","인감증명"]): return "🏢 법인서류"
    if any(k in fn for k in ["보고서","월간","자금수지","매출"]): return "📊 내부보고"
    if any(k in fn for k in ["공문","서약서","신청서","확약서","동의서"]): return "📋 공문/신청"
    if any(k in fn for k in ["영수증","수령","청구","약정금"]): return "💰 영수/청구"
    if any(k in fn for k in ["임대차","전대차"]): return "🏠 임대차"
    return "📄 기타"

def ir(d,days=7):
    try: return (datetime.now()-datetime.strptime(d,"%Y-%m-%d")).days<=days
    except: return False

def nb(d): return ' <span class="nb">NEW</span>' if ir(d) else ""
def ra(d): return ' data-recent="true"' if ir(d) else ""
def ub(u): return f'<span class="badge {u}">{u}</span>' if u else ""

def generate_html(lawsuit_dict,ids):
    try: pdfs=sorted([f for f in os.listdir(SCAN_FOLDER) if f.lower().endswith('.pdf')])
    except: pdfs=[]
    now_str=datetime.now().strftime("%Y-%m-%d %H:%M")
    all_docs=[]
    for f in pdfs:
        fid=ids.get(f,""); link=f"https://drive.google.com/file/d/{fid}/view" if fid else FOLDER_URL
        date=get_date(f)
        all_docs.append({"fname":f,"date":date,"fid":fid,"link":link,"company":get_co(f),"doctype":get_dt(f),"tags":get_tags(f),"is_lawsuit":f in lawsuit_dict,"urgency":lawsuit_dict.get(f,("","","","","일반"))[4],"info":lawsuit_dict.get(f)})
    all_docs.sort(key=lambda x:x["date"],reverse=True)
    ld=[d for d in all_docs if d["is_lawsuit"]]; wd=[d for d in all_docs if not d["is_lawsuit"]]
    un=sum(1 for d in ld if d["urgency"]=="긴급"); cn=sum(1 for d in ld if d["urgency"]=="주의")
    rn=sum(1 for d in all_docs if ir(d["date"])); ln=sum(1 for d in all_docs if d["fid"])
    coc={n:sum(1 for d in all_docs if n in d["tags"]) for n,_ in COMPANIES}

    lr="".join(f"""<tr{ra(d['date'])}><td class="c">{i+1}</td><td class="c dc">{d['date']}{nb(d['date'])}</td><td>{d['info'][0]}</td><td>{d['info'][1]}</td><td class="c">{d['info'][2]}</td><td class="c">{d['info'][3]}</td><td class="c">{ub(d['urgency'])}</td><td><a href="{d['link']}" target="_blank" class="fl">📄 {d['fname']}</a></td></tr>\n""" for i,d in enumerate(ld))
    wr="".join(f"""<tr{ra(d['date'])}><td class="c">{i+1}</td><td class="c dc">{d['date']}{nb(d['date'])}</td><td class="c">{d['company']}</td><td class="c">{d['doctype']}</td><td><a href="{d['link']}" target="_blank" class="fl">📄 {d['fname']}</a></td></tr>\n""" for i,d in enumerate(wd))

    ct=""; cp=""
    for cname,_ in COMPANIES:
        cnt=coc[cname]
        if cnt==0: continue
        short=cname.replace("플래닛네이처미디어","플래닛").replace("에스티엔미디어","STN미디어").replace("에스티엔뉴스","STN뉴스").replace("에스티엔","STN")
        ct+=f'<button class="tab co-tab" id="tab-co-{cname}" onclick="showTab(\'co-{cname}\',this)">{short} ({cnt})</button>\n'
        docs=[d for d in all_docs if cname in d["tags"]]
        rows="".join(f"""<tr{ra(d['date'])}><td class="c">{i+1}</td><td class="c dc">{d['date']}{nb(d['date'])}</td><td class="c">{d['doctype']}{' '+ub('주의') if d['is_lawsuit'] else ''}</td><td><a href="{d['link']}" target="_blank" class="fl">📄 {d['fname']}</a></td></tr>\n""" for i,d in enumerate(docs))
        cp+=f'<div id="panel-co-{cname}" class="panel"><table class="co-tbl"><thead><tr><th>No</th><th>스캔일자</th><th>문서종류</th><th>파일 열기</th></tr></thead><tbody>{rows}</tbody></table></div>\n'

    html=f"""<!DOCTYPE html><html lang="ko"><head><meta charset="UTF-8"><meta http-equiv="refresh" content="30"><title>스캔관리대장</title>
<style>*{{box-sizing:border-box;margin:0;padding:0}}body{{font-family:"맑은 고딕",sans-serif;background:#f0f2f5;padding:20px;font-size:13px}}h1{{color:#1F4E79;font-size:22px;margin-bottom:14px}}.summary{{display:flex;gap:10px;margin-bottom:14px;flex-wrap:wrap}}.card{{background:white;border-radius:10px;padding:12px 18px;box-shadow:0 2px 6px rgba(0,0,0,.1);text-align:center;min-width:80px}}.card .num{{font-size:26px;font-weight:bold;color:#1F4E79}}.card .label{{font-size:11px;color:#666;margin-top:3px}}.card.red .num{{color:#e53935}}.card.orange .num{{color:#FF9900}}.card.green .num{{color:#2e7d32}}.card.blue .num{{color:#1565c0}}.toolbar{{display:flex;gap:8px;margin-bottom:10px;align-items:center;flex-wrap:wrap}}.toolbar input{{padding:7px 12px;border:1px solid #ccc;border-radius:6px;font-size:13px;width:260px;font-family:"맑은 고딕"}}.toolbar select{{padding:7px 9px;border:1px solid #ccc;border-radius:6px;font-size:13px;font-family:"맑은 고딕"}}.btn{{padding:7px 14px;border:none;border-radius:6px;cursor:pointer;font-family:"맑은 고딕";font-size:13px;font-weight:bold}}.btn-reset{{background:#6c757d;color:white}}.btn-recent{{background:#2e7d32;color:white}}.btn-recent .cnt{{background:white;color:#2e7d32;border-radius:10px;padding:1px 6px;font-size:11px;margin-left:4px}}.btn-drive{{background:#4285f4;color:white;text-decoration:none;display:inline-flex;align-items:center}}.btn-refresh{{background:#1F4E79;color:white;margin-left:auto}}.btn:hover{{opacity:.85}}.tabs{{display:flex;gap:3px;flex-wrap:wrap;margin-bottom:0}}.tab{{padding:9px 16px;background:#ddd;border-radius:8px 8px 0 0;cursor:pointer;font-weight:bold;font-size:12px;border:none;font-family:"맑은 고딕";white-space:nowrap}}.tab.active{{background:#1F4E79;color:white}}.tab.green.active{{background:#375623}}.tab.co-tab.active{{background:#5c3d91;color:white}}.panel{{display:none}}.panel.active{{display:block}}table{{width:100%;border-collapse:collapse;background:white;box-shadow:0 2px 6px rgba(0,0,0,.1)}}th{{padding:9px 11px;text-align:center;border-bottom:2px solid #dee2e6;white-space:nowrap}}.lawsuit-tbl th{{background:#1F4E79;color:white}}.work-tbl th{{background:#375623;color:white}}.co-tbl th{{background:#5c3d91;color:white}}td{{padding:8px 11px;border-bottom:1px solid #e9ecef;vertical-align:middle}}tr:nth-child(even) td{{background:#fafafa}}tr:hover td{{background:#e8f4fd!important}}tr[data-recent="true"] td{{background:#f0fff4}}tr[data-recent="true"]:hover td{{background:#d4edda!important}}.c{{text-align:center}}.dc{{white-space:nowrap;font-weight:500}}a.fl{{color:#0563C1;text-decoration:none}}a.fl:hover{{text-decoration:underline}}.badge{{display:inline-block;padding:2px 9px;border-radius:10px;font-size:11px;font-weight:bold}}.긴급{{background:#e53935;color:white}}.주의{{background:#FF9900;color:white}}.일반{{background:#6c757d;color:white}}.nb{{display:inline-block;background:#2e7d32;color:white;font-size:10px;padding:1px 5px;border-radius:7px;margin-left:4px;font-weight:bold}}.hidden{{display:none!important}}.rbb{{background:#d4edda;border:1px solid #28a745;border-radius:6px;padding:7px 12px;margin-bottom:8px;color:#155724;font-size:12px;display:none}}.rbb.show{{display:block}}.updated{{color:#999;font-size:11px;margin-top:10px;text-align:right}}</style></head>
<body><h1>📂 스캔관리대장 <span style="font-size:12px;color:#888;font-weight:normal">🤖 AI 자동분석 | 30초 자동갱신</span></h1>
<div class="summary"><div class="card"><div class="num">{len(ld)}</div><div class="label">소송문서</div></div><div class="card red"><div class="num">{un}</div><div class="label">🔴 긴급</div></div><div class="card orange"><div class="num">{cn}</div><div class="label">🟡 주의</div></div><div class="card"><div class="num">{len(wd)}</div><div class="label">업무문서</div></div><div class="card green"><div class="num">{rn}</div><div class="label">📅 최근7일</div></div><div class="card blue"><div class="num">{ln}</div><div class="label">☁️ 드라이브</div></div><div class="card"><div class="num">{len(pdfs)}</div><div class="label">전체</div></div></div>
<div class="toolbar"><input type="text" id="si" placeholder="🔍 파일명, 회사명, 사건명..." oninput="doSearch()"><select id="sc" onchange="doSearch()"><option value="all">전체탭</option><option value="lawsuit">소송만</option><option value="work">업무만</option></select><button class="btn btn-reset" onclick="clearAll()">초기화</button><button class="btn btn-recent" id="rb" onclick="toggleRecent()">📅 최근 7일 <span class="cnt">{rn}</span></button><a href="{FOLDER_URL}" target="_blank" class="btn btn-drive">☁️ 드라이브</a><span id="ci" style="color:#666;font-size:12px"></span><button class="btn btn-refresh" onclick="location.reload()">🔄 새로고침</button></div>
<div id="rbb" class="rbb">📅 최근 7일 문서만 표시 중 &nbsp;<a href="#" onclick="clearAll();return false;">전체 보기</a></div>
<div class="tabs"><button class="tab active" id="tab-lawsuit" onclick="showTab('lawsuit',this)">⚖️ 소송관리 ({len(ld)})</button><button class="tab green" id="tab-work" onclick="showTab('work',this)">📋 업무문서 ({len(wd)})</button>{ct}</div>
<div id="panel-lawsuit" class="panel active"><table class="lawsuit-tbl"><thead><tr><th>No</th><th>스캔일자</th><th>회사(기관)명</th><th>사건명</th><th>사건번호</th><th>문서종류</th><th>긴급도</th><th>파일 열기</th></tr></thead><tbody id="lb">{lr}</tbody></table></div>
<div id="panel-work" class="panel"><table class="work-tbl"><thead><tr><th>No</th><th>스캔일자</th><th>회사(기관)명</th><th>문서종류</th><th>파일 열기</th></tr></thead><tbody id="wb">{wr}</tbody></table></div>
{cp}
<p class="updated">※ 갱신:{now_str} | 🤖 AI자동분석 | ☁️ 구글드라이브 | 드라이브링크:{ln}/{len(pdfs)}</p>
<script>let rm=false;function showTab(n,btn){{document.querySelectorAll('.panel').forEach(p=>p.classList.remove('active'));document.querySelectorAll('.tab').forEach(t=>t.classList.remove('active'));document.getElementById('panel-'+n).classList.add('active');btn.classList.add('active');applyF();}}function toggleRecent(){{rm=!rm;document.getElementById('rb').style.background=rm?'#155724':'#2e7d32';document.getElementById('rbb').classList.toggle('show',rm);applyF();}}function doSearch(){{rm=false;document.getElementById('rb').style.background='#2e7d32';document.getElementById('rbb').classList.remove('show');applyF();}}function applyF(){{const q=document.getElementById('si').value.toLowerCase().trim();let v=0;document.querySelectorAll('tbody tr').forEach(tr=>{{const ok=(!q||tr.textContent.toLowerCase().includes(q))&&(!rm||tr.dataset.recent==='true');tr.classList.toggle('hidden',!ok);if(ok)v++;}});document.getElementById('ci').textContent=(q||rm)?`표시: ${{v}}건`:''}}function clearAll(){{document.getElementById('si').value='';document.getElementById('sc').value='all';rm=false;document.getElementById('rb').style.background='#2e7d32';document.getElementById('rbb').classList.remove('show');document.getElementById('ci').textContent='';document.querySelectorAll('tr').forEach(tr=>tr.classList.remove('hidden'));}}</script></body></html>"""

    with open(OUTPUT_HTML,"w",encoding="utf-8") as f: f.write(html)
    log.info(f"HTML갱신완료({len(pdfs)}개,소송:{len(ld)},드라이브:{ln})")

def process_pdf(pdf_path,done_db):
    fname=os.path.basename(pdf_path)
    if fname in done_db: return False
    log.info(f"▶ 처리: {fname}")
    prev=-1
    for _ in range(20):
        try:
            sz=os.path.getsize(pdf_path)
            if sz==prev and sz>0: break
            prev=sz; time.sleep(1)
        except: return False
    text=extract_text(pdf_path); img_b64=None
    if not text.strip(): img_b64=pdf_to_b64(pdf_path)
    if not text and not img_b64: log.warning(f"추출실패:{fname}"); return False
    meta=analyze_pdf(fname,text,img_b64)
    log.info(f"  [{meta['type']}] {meta.get('case_name','')} [{meta.get('urgency','')}]")
    new_path=safe_rename(pdf_path,meta.get("new_filename",Path(fname).stem))
    new_fname=os.path.basename(new_path)
    if meta["type"]=="소송": update_lawsuit_excel(meta,new_fname)
    done_db[new_fname]={"original":fname,"type":meta["type"],"processed_at":datetime.now().isoformat()}
    save_done(done_db)
    return True

def main():
    print("="*55)
    print("  📂 스캔 자동화 통합 관리 시스템 v1.0")
    print("="*55)
    if not load_keys():
        print("\n❌ api_key.txt 확인하세요\n형식:\nANTHROPIC=sk-ant-api03-...\nGDRIVE=AIza...")
        input("\nEnter 눌러 종료..."); return
    log.info("시스템 시작")
    done_db=load_done()
    ids=load_drive_ids()
    pdfs=[f for f in os.listdir(SCAN_FOLDER) if f.lower().endswith('.pdf')]
    unprocessed=[f for f in pdfs if f not in done_db]
    if unprocessed:
        log.info(f"미처리파일 {len(unprocessed)}개 처리")
        for fname in unprocessed:
            process_pdf(os.path.join(SCAN_FOLDER,fname),done_db); time.sleep(0.5)
        ids=load_drive_ids(force=True)
    generate_html(load_lawsuit_dict(),ids)
    log.info(f"폴더감시시작:{SCAN_FOLDER}")
    print(f"\n✅ 실행 중 - 스캔 폴더에 PDF 추가하면 자동 처리됩니다")
    print("   종료: Ctrl+C\n")
    prev_files=set(os.listdir(SCAN_FOLDER)); html_timer=time.time()
    while True:
        try:
            cur=set(os.listdir(SCAN_FOLDER)); new=cur-prev_files
            if new:
                for fname in new:
                    if not fname.lower().endswith('.pdf'): continue
                    time.sleep(2)
                    if process_pdf(os.path.join(SCAN_FOLDER,fname),done_db):
                        ids=load_drive_ids(force=True); generate_html(load_lawsuit_dict(),ids)
                prev_files=set(os.listdir(SCAN_FOLDER))
            if time.time()-html_timer>1800:
                ids=load_drive_ids(force=True); generate_html(load_lawsuit_dict(),ids); html_timer=time.time()
            time.sleep(3)
        except KeyboardInterrupt: log.info("종료"); break
        except Exception as e: log.error(f"오류:{e}"); time.sleep(5)

if __name__=="__main__":
    main()

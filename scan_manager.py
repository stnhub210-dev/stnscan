"""
스캔 자동 관리 시스템
- PDF 내용 분석 → 파일명 자동 변경
- 소송 / 업무문서 엑셀 자동 정리
- 폴더 감시 → 신규 파일 추가 시 자동 업데이트
"""

import os
import sys
import time
import shutil
import hashlib
import json
import logging
import re
from datetime import datetime
from pathlib import Path
import threading

# ─────────────────────────────────────────────
# 설정 (사용자 환경에 맞게 수정)
# ─────────────────────────────────────────────
import os as _os

SCAN_FOLDER = _os.path.join(_os.path.expanduser("~"), "Desktop", "스캔")
EXCEL_PATH  = _os.path.join(_os.path.expanduser("~"), "Desktop", "스캔관리대장.xlsx")
LOG_PATH    = _os.path.join(_os.path.expanduser("~"), "Desktop", "scan_manager.log")
PROCESSED_DB = _os.path.join(_os.path.expanduser("~"), "Desktop", "scan_processed.json")

# Anthropic API 키 (환경변수 ANTHROPIC_API_KEY 또는 아래 직접 입력)
ANTHROPIC_API_KEY = _os.environ.get("ANTHROPIC_API_KEY", "")

# ─────────────────────────────────────────────
# 로깅 설정
# ─────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(LOG_PATH, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger(__name__)


# ─────────────────────────────────────────────
# 처리 기록 DB (중복 처리 방지)
# ─────────────────────────────────────────────
def load_processed():
    if os.path.exists(PROCESSED_DB):
        with open(PROCESSED_DB, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def save_processed(db: dict):
    with open(PROCESSED_DB, "w", encoding="utf-8") as f:
        json.dump(db, f, ensure_ascii=False, indent=2)

def file_hash(path: str) -> str:
    h = hashlib.md5()
    with open(path, "rb") as f:
        h.update(f.read(8192))
    return h.hexdigest()


# ─────────────────────────────────────────────
# PDF 텍스트 추출
# ─────────────────────────────────────────────
def extract_pdf_text(pdf_path: str, max_pages: int = 5) -> str:
    text_parts = []

    # 방법 1: pdfplumber (텍스트 PDF)
    try:
        import pdfplumber
        with pdfplumber.open(pdf_path) as pdf:
            for i, page in enumerate(pdf.pages[:max_pages]):
                t = page.extract_text() or ""
                if t.strip():
                    text_parts.append(t)
        if text_parts:
            return "\n".join(text_parts)[:4000]
    except Exception as e:
        log.debug(f"pdfplumber 실패: {e}")

    # 방법 2: OCR (스캔 이미지 PDF)
    try:
        from pdf2image import convert_from_path
        import pytesseract

        # Windows Tesseract 경로 (설치 위치에 따라 조정)
        if sys.platform == "win32":
            pytesseract.pytesseract.tesseract_cmd = (
                r"C:\Program Files\Tesseract-OCR\tesseract.exe"
            )

        images = convert_from_path(pdf_path, first_page=1, last_page=min(max_pages, 3), dpi=150)
        for img in images:
            t = pytesseract.image_to_string(img, lang="kor+eng")
            if t.strip():
                text_parts.append(t)
        if text_parts:
            return "\n".join(text_parts)[:4000]
    except Exception as e:
        log.debug(f"OCR 실패: {e}")

    return ""


# ─────────────────────────────────────────────
# Claude API 분석
# ─────────────────────────────────────────────
def analyze_with_claude(text: str, filename: str) -> dict:
    """
    Claude API로 PDF 내용 분석 → 분류 및 메타데이터 반환
    반환 형식:
    {
      "type": "소송" | "업무",
      "new_filename": "새파일명(확장자제외)",
      "company": "회사명",
      "case_name": "사건명",
      "case_number": "사건번호",
      "cert_number": "인증번호",
      "deadline": "마감일(YYYY-MM-DD)",
      "urgency": "긴급" | "주의" | "일반",
      "summary": "요약",
      "scan_date": "스캔날짜(YYYY-MM-DD)"
    }
    """
    try:
        import anthropic
        client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)

        today = datetime.now().strftime("%Y-%m-%d")
        prompt = f"""아래는 스캔된 PDF 문서의 텍스트입니다. 내용을 분석하여 JSON으로만 응답하세요.

파일명: {filename}
오늘 날짜: {today}
내용:
{text[:3000]}

다음 JSON 형식으로만 응답 (다른 텍스트 없이):
{{
  "type": "소송 또는 업무 중 하나",
  "new_filename": "내용에 맞는 한국어 파일명(날짜_회사명_문서종류 형식, 특수문자/슬래시 제외, 50자 이내)",
  "company": "관련 회사명 또는 기관명",
  "case_name": "사건명 또는 문서명",
  "case_number": "사건번호(없으면 빈문자열)",
  "cert_number": "인증번호 또는 접수번호(없으면 빈문자열)",
  "deadline": "마감일 또는 기한(YYYY-MM-DD 형식, 없으면 빈문자열)",
  "urgency": "마감 7일 이내면 긴급, 30일 이내면 주의, 그 외 일반",
  "summary": "문서 핵심 내용 2~3줄 요약",
  "scan_date": "문서에 기재된 날짜 또는 오늘날짜(YYYY-MM-DD)"
}}

판단 기준:
- type=소송: 소장, 답변서, 결정문, 통지서, 법원서류, 채권, 회생, 소송 관련
- type=업무: 계약서, 제안서, 기획서, 보고서, 견적서, 공문, 기타 업무문서"""

        resp = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=800,
            messages=[{"role": "user", "content": prompt}],
        )
        raw = resp.content[0].text.strip()
        # JSON 펜스 제거
        raw = re.sub(r"```json|```", "", raw).strip()
        data = json.loads(raw)
        return data

    except Exception as e:
        log.error(f"Claude 분석 오류: {e}")
        return {
            "type": "업무",
            "new_filename": Path(filename).stem,
            "company": "",
            "case_name": filename,
            "case_number": "",
            "cert_number": "",
            "deadline": "",
            "urgency": "일반",
            "summary": "분석 실패",
            "scan_date": datetime.now().strftime("%Y-%m-%d"),
        }


# ─────────────────────────────────────────────
# 파일명 변경
# ─────────────────────────────────────────────
def safe_rename(original_path: str, new_basename: str) -> str:
    """충돌 없이 파일명 변경, 새 경로 반환"""
    folder = os.path.dirname(original_path)
    ext = Path(original_path).suffix
    # 파일명 안전화: 특수문자 제거
    safe = re.sub(r'[\\/:*?"<>|]', "_", new_basename)
    new_path = os.path.join(folder, safe + ext)

    # 동일명 충돌 처리
    counter = 1
    while os.path.exists(new_path) and new_path != original_path:
        new_path = os.path.join(folder, f"{safe}_{counter}{ext}")
        counter += 1

    if new_path != original_path:
        os.rename(original_path, new_path)
        log.info(f"파일명 변경: {os.path.basename(original_path)} → {os.path.basename(new_path)}")

    return new_path


# ─────────────────────────────────────────────
# 엑셀 업데이트
# ─────────────────────────────────────────────
def get_or_create_workbook():
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter

    if os.path.exists(EXCEL_PATH):
        wb = load_workbook(EXCEL_PATH)
    else:
        wb = Workbook()
        # 기본 시트 제거
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    # ── 소송 시트 ──
    if "소송관리" not in wb.sheetnames:
        ws = wb.create_sheet("소송관리", 0)
        headers = ["No", "회사(기관)명", "사건명", "사건번호", "인증/접수번호",
                   "마감일", "D-Day", "긴급도", "문서요약", "파일명", "등록일시"]
        _style_header(ws, headers)
        _set_col_widths(ws, [5, 20, 25, 20, 20, 13, 8, 8, 40, 35, 18])
    
    # ── 업무문서 시트 ──
    if "업무문서" not in wb.sheetnames:
        ws = wb.create_sheet("업무문서", 1)
        headers = ["No", "스캔날짜", "회사(기관)명", "문서명", "문서요약", "파일명", "등록일시"]
        _style_header(ws, headers)
        _set_col_widths(ws, [5, 13, 20, 30, 45, 40, 18])

    return wb


def _style_header(ws, headers):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


def _set_col_widths(ws, widths):
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _row_border(ws, row_idx, col_count):
    from openpyxl.styles import Border, Side, Alignment
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.border = border
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def _dday_color(ws, row, col, deadline_str):
    """D-Day 계산 + 색상 강조"""
    from openpyxl.styles import PatternFill, Font
    if not deadline_str:
        ws.cell(row=row, column=col, value="-")
        return
    try:
        dl = datetime.strptime(deadline_str, "%Y-%m-%d")
        diff = (dl - datetime.now()).days
        label = f"D-{abs(diff)}" if diff > 0 else ("D-Day" if diff == 0 else f"D+{abs(diff)}")
        cell = ws.cell(row=row, column=col, value=label)
        if diff <= 7:
            cell.fill = PatternFill(fill_type="solid", fgColor="FF0000")
            cell.font = Font(color="FFFFFF", bold=True)
        elif diff <= 30:
            cell.fill = PatternFill(fill_type="solid", fgColor="FF9900")
            cell.font = Font(bold=True)
    except Exception:
        ws.cell(row=row, column=col, value="-")


def _urgency_color(ws, row, col, urgency):
    from openpyxl.styles import PatternFill, Font
    cell = ws.cell(row=row, column=col, value=urgency)
    if urgency == "긴급":
        cell.fill = PatternFill(fill_type="solid", fgColor="FF0000")
        cell.font = Font(color="FFFFFF", bold=True)
    elif urgency == "주의":
        cell.fill = PatternFill(fill_type="solid", fgColor="FFCC00")
        cell.font = Font(bold=True)


def add_to_excel(meta: dict, final_filename: str):
    wb = get_or_create_workbook()
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")

    if meta.get("type") == "소송":
        ws = wb["소송관리"]
        next_row = ws.max_row + 1
        row_no = next_row - 1

        ws.cell(next_row, 1, row_no)
        ws.cell(next_row, 2, meta.get("company", ""))
        ws.cell(next_row, 3, meta.get("case_name", ""))
        ws.cell(next_row, 4, meta.get("case_number", ""))
        ws.cell(next_row, 5, meta.get("cert_number", ""))
        ws.cell(next_row, 6, meta.get("deadline", ""))
        _dday_color(ws, next_row, 7, meta.get("deadline", ""))
        _urgency_color(ws, next_row, 8, meta.get("urgency", "일반"))
        ws.cell(next_row, 9, meta.get("summary", ""))
        ws.cell(next_row, 10, final_filename)
        ws.cell(next_row, 11, now_str)
        _row_border(ws, next_row, 11)
        ws.row_dimensions[next_row].height = 35

    else:
        ws = wb["업무문서"]
        next_row = ws.max_row + 1
        row_no = next_row - 1

        ws.cell(next_row, 1, row_no)
        ws.cell(next_row, 2, meta.get("scan_date", ""))
        ws.cell(next_row, 3, meta.get("company", ""))
        ws.cell(next_row, 4, meta.get("case_name", ""))
        ws.cell(next_row, 5, meta.get("summary", ""))
        ws.cell(next_row, 6, final_filename)
        ws.cell(next_row, 7, now_str)
        _row_border(ws, next_row, 7)
        ws.row_dimensions[next_row].height = 35

    wb.save(EXCEL_PATH)
    log.info(f"엑셀 업데이트 완료: [{meta.get('type')}] {final_filename}")


# ─────────────────────────────────────────────
# 핵심 처리 함수
# ─────────────────────────────────────────────
def process_pdf(pdf_path: str, processed_db: dict):
    """단일 PDF 처리 파이프라인"""
    try:
        fhash = file_hash(pdf_path)
        if fhash in processed_db:
            log.info(f"이미 처리된 파일 건너뜀: {os.path.basename(pdf_path)}")
            return

        log.info(f"처리 시작: {os.path.basename(pdf_path)}")

        # 1. 텍스트 추출
        text = extract_pdf_text(pdf_path)
        if not text.strip():
            log.warning(f"텍스트 추출 실패 (빈 내용): {os.path.basename(pdf_path)}")
            text = f"파일명: {os.path.basename(pdf_path)}"

        # 2. Claude 분석
        meta = analyze_with_claude(text, os.path.basename(pdf_path))

        # 3. 파일명 변경
        new_path = safe_rename(pdf_path, meta.get("new_filename", Path(pdf_path).stem))
        final_filename = os.path.basename(new_path)

        # 4. 엑셀 업데이트
        add_to_excel(meta, final_filename)

        # 5. 처리 완료 기록
        processed_db[fhash] = {
            "original": os.path.basename(pdf_path),
            "renamed": final_filename,
            "processed_at": datetime.now().isoformat(),
            "type": meta.get("type"),
        }
        save_processed(processed_db)

    except Exception as e:
        log.error(f"처리 오류 [{os.path.basename(pdf_path)}]: {e}", exc_info=True)


# ─────────────────────────────────────────────
# 폴더 감시 (watchdog)
# ─────────────────────────────────────────────
def start_watcher(processed_db: dict):
    from watchdog.observers import Observer
    from watchdog.events import FileSystemEventHandler

    class PDFHandler(FileSystemEventHandler):
        def __init__(self):
            self._pending = set()
            self._lock = threading.Lock()

        def on_created(self, event):
            if not event.is_directory and event.src_path.lower().endswith(".pdf"):
                # 파일 쓰기 완료 대기 후 처리
                threading.Thread(
                    target=self._delayed_process,
                    args=(event.src_path,),
                    daemon=True,
                ).start()

        def _delayed_process(self, path):
            # 스캐너가 파일 쓰기 완료할 때까지 대기
            time.sleep(3)
            prev_size = -1
            for _ in range(30):
                try:
                    size = os.path.getsize(path)
                    if size == prev_size and size > 0:
                        break
                    prev_size = size
                    time.sleep(1)
                except FileNotFoundError:
                    return

            with self._lock:
                if path not in self._pending:
                    self._pending.add(path)
                    process_pdf(path, processed_db)
                    self._pending.discard(path)

    observer = Observer()
    handler = PDFHandler()
    observer.schedule(handler, SCAN_FOLDER, recursive=False)
    observer.start()
    log.info(f"📂 폴더 감시 시작: {SCAN_FOLDER}")
    return observer


# ─────────────────────────────────────────────
# 기존 파일 일괄 처리
# ─────────────────────────────────────────────
def process_existing(processed_db: dict):
    if not os.path.exists(SCAN_FOLDER):
        log.warning(f"스캔 폴더 없음, 생성: {SCAN_FOLDER}")
        os.makedirs(SCAN_FOLDER, exist_ok=True)
        return

    pdfs = sorted(Path(SCAN_FOLDER).glob("*.pdf"))
    if not pdfs:
        log.info("처리할 PDF 파일 없음")
        return

    log.info(f"기존 PDF {len(pdfs)}개 일괄 처리 시작")
    for p in pdfs:
        process_pdf(str(p), processed_db)
        time.sleep(0.5)  # API 속도 제한


# ─────────────────────────────────────────────
# 메인
# ─────────────────────────────────────────────
def main():
    print("=" * 60)
    print("  스캔 자동 관리 시스템 v1.0")
    print("=" * 60)

    if not ANTHROPIC_API_KEY:
        print("\n❌ ANTHROPIC_API_KEY 환경변수를 설정해주세요.")
        print("   예) set ANTHROPIC_API_KEY=sk-ant-...")
        sys.exit(1)

    log.info(f"스캔: {SCAN_FOLDER}")
    log.info(f"엑셀경로: {EXCEL_PATH}")

    processed_db = load_processed()

    # 기존 파일 처리
    process_existing(processed_db)

    # 폴더 감시 시작
    observer = start_watcher(processed_db)

    print(f"\n✅ 실행 중. 새 PDF를 [{SCAN_FOLDER}]에 추가하면 자동 처리됩니다.")
    print("   종료하려면 Ctrl+C\n")

    try:
        while True:
            time.sleep(5)
    except KeyboardInterrupt:
        observer.stop()
        log.info("프로그램 종료")

    observer.join()


if __name__ == "__main__":
    main()
